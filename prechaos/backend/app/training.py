from __future__ import annotations

import json
from dataclasses import dataclass
from pathlib import Path
from typing import Iterable

import numpy as np
from openpyxl import load_workbook
from sklearn.metrics import accuracy_score, confusion_matrix, f1_score, precision_score, recall_score
from sklearn.model_selection import train_test_split
from sklearn.preprocessing import StandardScaler

from .config import FEATURE_NAMES, LIVE_TRAINING_META_PATH, MODEL_ROOT, SCALER_PATH, TRAINED_MODEL_PATH, WINDOW_SIZE
from .feature_engineering import build_feature_matrix, sanitize_events, validate_feature_vector
from .model import build_classifier
from .security import safe_joblib_dump, safe_write_text, write_artifact_manifest


@dataclass
class TrainingArtifacts:
    model_path: str
    scaler_path: str
    metrics: dict


def _load_dataset(dataset_path: Path) -> list[dict]:
    if dataset_path.suffix.lower() == ".json":
        return json.loads(dataset_path.read_text(encoding="utf-8"))
    if dataset_path.suffix.lower() == ".jsonl":
        records: list[dict] = []
        for line in dataset_path.read_text(encoding="utf-8").splitlines():
            if line.strip():
                records.append(json.loads(line))
        return records
    if dataset_path.suffix.lower() in {".xlsx", ".xlsm"}:
        workbook = load_workbook(dataset_path, read_only=True, data_only=True)
        sheet = workbook[workbook.sheetnames[0]]
        rows = sheet.iter_rows(values_only=True)
        header = [str(value) for value in next(rows)]
        records: list[dict] = []
        for row in rows:
            if row and row[0]:
                records.append({header[index]: row[index] for index in range(len(header))})
        return records
    raise ValueError("Supported dataset formats are JSON, JSONL, and XLSX.")


def _engineer_features(row: dict) -> np.ndarray:
    if "features" in row and row["features"] is not None:
        return np.asarray(validate_feature_vector(row["features"]), dtype=np.float32)
    if "events" in row and row["events"] is not None:
        feature_matrix = build_feature_matrix(sanitize_events(row["events"]))
        return np.asarray(feature_matrix[-1], dtype=np.float32)
    if "subject" in row:
        hold_values = np.array(
            [float(value) for key, value in row.items() if str(key).startswith("H.") and value is not None],
            dtype=np.float32,
        )
        dd_values = np.array(
            [float(value) for key, value in row.items() if str(key).startswith("DD.") and value is not None],
            dtype=np.float32,
        )
        ud_values = np.array(
            [float(value) for key, value in row.items() if str(key).startswith("UD.") and value is not None],
            dtype=np.float32,
        )
        hold_mean = float(np.mean(hold_values)) if hold_values.size else 0.12
        dd_mean = float(np.mean(dd_values)) if dd_values.size else 0.18
        ud_mean = float(np.mean(ud_values)) if ud_values.size else 0.09
        typing_speed = 1.0 / max(hold_mean, 1e-4)
        pause_time = max(dd_mean, ud_mean)
        variation = float(np.std(np.concatenate([hold_values, dd_values, ud_values]))) if (
            hold_values.size or dd_values.size or ud_values.size
        ) else 0.0
        error_score = float(np.mean(np.abs(dd_values - ud_values[: dd_values.size]))) if (
            dd_values.size and ud_values.size
        ) else 0.0
        idle_time = max(float(row.get("sessionIndex", 1)) - 1.0, 0.0) * 0.15
        mouse_speed = 0.0
        tab_switch = max(float(row.get("rep", 1)) - 1.0, 0.0) / 50.0
        session_duration = float(row.get("rep", 1)) / 5.0
        explicit_fatigue = row.get("fatigue_score")
        fatigue_score = (
            float(explicit_fatigue)
            if explicit_fatigue not in (None, "")
            else float(np.clip(variation / max(hold_mean, 1e-4), 0.0, 1.0))
        )
        return np.array(
            [
                typing_speed,
                pause_time,
                variation,
                error_score,
                idle_time,
                mouse_speed,
                tab_switch,
                session_duration,
                fatigue_score,
            ],
            dtype=np.float32,
        )

    hold = float(row.get("hold_time", row.get("typing_speed", 150.0)))
    dd = float(row.get("dd_latency", row.get("pause_time", 80.0)))
    ud = float(row.get("ud_latency", row.get("variation", 60.0)))
    deviation = float(row.get("deviation", abs(dd - ud)))
    idle_time = float(row.get("idle_time", 0.0))
    mouse_speed = float(row.get("mouse_movement_speed", 0.0))
    tab_switch = float(row.get("tab_switch_frequency", 0.0))
    session_duration = float(row.get("session_duration", 0.0))
    fatigue_score = float(row.get("fatigue_score", 0.0))
    typing_speed = 1000.0 / max(hold, 1.0)
    pause_time = max(dd, ud) / 1000.0
    variation = np.std([hold, dd, ud]) / 1000.0
    error_score = deviation / max(np.mean([hold, dd, ud]), 1.0)
    return np.array(
        [
            typing_speed,
            pause_time,
            variation,
            error_score,
            idle_time,
            mouse_speed,
            tab_switch,
            session_duration,
            fatigue_score,
        ],
        dtype=np.float32,
    )


def _sort_key(record: dict) -> tuple[float, str]:
    try:
        timestamp = float(record.get("timestamp") or 0.0)
    except (TypeError, ValueError):
        timestamp = 0.0
    return (timestamp, str(record.get("session_id", "")))


def _build_labeled_samples(records: Iterable[dict]) -> tuple[np.ndarray, np.ndarray, np.ndarray]:
    by_subject: dict[str, list[np.ndarray]] = {}
    by_session: dict[str, list[np.ndarray]] = {}
    for record in sorted(records, key=_sort_key):
        if "prediction" in record:
            raise ValueError("Dataset contains prediction fields. Refusing to train on contaminated records.")
        subject_id = str(record.get("user_id", record.get("subject", "unknown")))
        session_id = str(record.get("session_id", record.get("sessionIndex", "0")))
        features = _engineer_features(record)
        by_subject.setdefault(subject_id, []).append(features)
        by_session.setdefault(f"{subject_id}:{session_id}", []).append(features)

    sample_vectors: list[np.ndarray] = []
    sample_labels: list[int] = []
    sample_sessions: list[str] = []
    session_samples: list[tuple[str, list[np.ndarray], float]] = []
    session_scores: list[float] = []

    for session_key, samples in by_session.items():
        if len(samples) < WINDOW_SIZE:
            continue
        matrix = np.stack(samples)
        subject_id = session_key.split(":", 1)[0]
        subject_baseline = np.stack(by_subject[subject_id]).mean(axis=0)

        windows: list[np.ndarray] = []
        window_scores: list[float] = []
        for start in range(0, len(matrix) - WINDOW_SIZE + 1):
            window = matrix[start : start + WINDOW_SIZE]
            windows.append(window[-1].astype(np.float32))
            instability = float(np.mean(np.linalg.norm(window - subject_baseline, axis=1)))
            window_scores.append(instability)

        if not window_scores:
            continue

        session_score = float(np.median(window_scores))
        session_samples.append((session_key, windows, session_score))
        session_scores.append(session_score)

    if not session_samples:
        raise ValueError(f"Dataset did not produce any valid sessions with at least {WINDOW_SIZE} samples.")

    threshold = float(np.median(session_scores))
    for session_key, windows, session_score in session_samples:
        session_label = 1 if session_score >= threshold else 0
        for window_vector in windows:
            sample_vectors.append(window_vector)
            sample_labels.append(session_label)
            sample_sessions.append(session_key)

    return (
        np.stack(sample_vectors),
        np.asarray(sample_labels, dtype=np.int32),
        np.asarray(sample_sessions, dtype=object),
    )


def _split_by_session(
    features: np.ndarray,
    labels: np.ndarray,
    session_ids: np.ndarray,
) -> tuple[np.ndarray, np.ndarray, np.ndarray, np.ndarray, np.ndarray, np.ndarray]:
    unique_sessions, first_indices = np.unique(session_ids, return_index=True)
    if len(unique_sessions) < 2:
        raise ValueError("At least two sessions are required to create a train/validation split.")
    session_labels = labels[first_indices]
    stratify = session_labels if len(np.unique(session_labels)) > 1 else None

    train_sessions, val_sessions = train_test_split(
        unique_sessions,
        test_size=0.2,
        random_state=42,
        stratify=stratify,
    )
    train_mask = np.isin(session_ids, train_sessions)
    val_mask = np.isin(session_ids, val_sessions)
    if not np.any(train_mask) or not np.any(val_mask):
        raise ValueError("Training split failed to produce both train and validation samples.")

    return (
        features[train_mask],
        features[val_mask],
        labels[train_mask],
        labels[val_mask],
        train_sessions,
        val_sessions,
    )


def train_from_dataset(
    dataset_file: str,
    epochs: int = 24,
    batch_size: int = 32,
    learning_rate: float = 1e-3,
) -> TrainingArtifacts:
    # Epochs, batch_size, and learning_rate remain accepted for API compatibility.
    _ = (epochs, batch_size, learning_rate)

    dataset_path = Path(dataset_file).expanduser().resolve()
    records = _load_dataset(dataset_path)
    features, labels, session_ids = _build_labeled_samples(records)
    X_train, X_val, y_train, y_val, train_sessions, val_sessions = _split_by_session(features, labels, session_ids)

    scaler = StandardScaler()
    X_train_scaled = scaler.fit_transform(X_train)
    X_val_scaled = scaler.transform(X_val)

    model = build_classifier()
    model.fit(X_train_scaled, y_train)

    probs = model.predict_proba(X_val_scaled)[:, 1]
    preds = (probs >= 0.5).astype(np.int32)
    metrics = {
        "accuracy": round(float(accuracy_score(y_val, preds)), 4),
        "precision": round(float(precision_score(y_val, preds, zero_division=0)), 4),
        "recall": round(float(recall_score(y_val, preds, zero_division=0)), 4),
        "f1": round(float(f1_score(y_val, preds, zero_division=0)), 4),
        "confusion_matrix": confusion_matrix(y_val, preds).tolist(),
        "train_samples": int(len(X_train)),
        "validation_samples": int(len(X_val)),
        "train_sessions": int(len(train_sessions)),
        "validation_sessions": int(len(val_sessions)),
        "model": "logistic_regression",
    }

    safe_joblib_dump(model, TRAINED_MODEL_PATH, managed_root=MODEL_ROOT)
    safe_joblib_dump(scaler, SCALER_PATH, managed_root=MODEL_ROOT)

    training_metadata = {
        "last_trained_at": str(np.datetime64("now")),
        "metrics": metrics,
        "dataset_file": dataset_path.name,
        "session_count": len(np.unique(session_ids)),
        "model_type": "logistic_regression",
        "feature_count": len(FEATURE_NAMES),
    }
    write_artifact_manifest(
        model_path=TRAINED_MODEL_PATH,
        scaler_path=SCALER_PATH,
        metadata=training_metadata,
    )
    safe_write_text(LIVE_TRAINING_META_PATH, json.dumps(training_metadata, indent=2), managed_root=LIVE_TRAINING_META_PATH.parent)

    return TrainingArtifacts(
        model_path=str(TRAINED_MODEL_PATH),
        scaler_path=str(SCALER_PATH),
        metrics=metrics,
    )
